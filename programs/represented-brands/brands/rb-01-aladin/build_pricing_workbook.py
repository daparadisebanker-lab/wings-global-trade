# Generates ALADIN_CUPO_PRICING_v1.xlsx — RB/01 cupo pricing model.
# Re-run after structural changes; values are edited in the workbook itself
# (INPUTS drivers), formulas recalculate live in Excel.
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\Muaaz\projects\wings-global-trade\programs\represented-brands\brands\rb-01-aladin\ALADIN_CUPO_PRICING_v1.xlsx"

SOL = '"S/" #,##0.00'
SOL0 = '"S/" #,##0'
USD = '"US$" #,##0'
PCT = '0.0%'
NUM = '#,##0'
NUM2 = '#,##0.00'

GREEN = "4C7012"   # aladin accent-ink
CREAM = "FCFCF7"
YELLOW = "FFF3B0"  # editable driver
GREY = "F2F2F2"

f_title = Font(bold=True, size=14, color=GREEN)
f_sec = Font(bold=True, size=11, color="FFFFFF")
f_hdr = Font(bold=True, size=10)
f_note = Font(italic=True, size=9, color="666666")
f_warn = Font(bold=True, size=9, color="A61B3A")
fill_sec = PatternFill("solid", fgColor=GREEN)
fill_edit = PatternFill("solid", fgColor=YELLOW)
fill_hdr = PatternFill("solid", fgColor=GREY)
thin = Side(style="thin", color="CCCCCC")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def sec(ws, row, text, span=3):
    ws.cell(row=row, column=1, value=text).font = f_sec
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = fill_sec

def driver(ws, row, label, value, fmt=NUM2, note=None):
    ws.cell(row=row, column=1, value=label)
    c = ws.cell(row=row, column=3, value=value)
    c.fill = fill_edit
    c.number_format = fmt
    c.border = box
    if note:
        ws.cell(row=row, column=4, value=note).font = f_note

def calc(ws, row, label, formula, fmt=NUM2, note=None):
    ws.cell(row=row, column=1, value=label)
    c = ws.cell(row=row, column=3, value=formula)
    c.number_format = fmt
    c.border = box
    if note:
        ws.cell(row=row, column=4, value=note).font = f_note

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w

# ───────────────────────────── LEEME ─────────────────────────────
ws = wb.active
ws.title = "LEEME"
ws["A1"] = "RB/01 ÁLADÍN — MODELO DE PRECIOS POR CUPO · v1"
ws["A1"].font = f_title
ws["A2"] = "2026-07-16 · Fuente de decisión para las compuertas §12 de GTM-CAMPAIGN.md"
ws["A2"].font = f_note
rows = [
    ("", ""),
    ("CÓMO USAR", "Edita solo las celdas amarillas (hoja INPUTS y tarifas en FLETE). Todo lo demás son fórmulas."),
    ("", ""),
    ("DECISIONES QUE SALEN DE ESTE LIBRO", ""),
    ("1 · Markup base Wings", "INPUTS C33 — define el precio de cupo publicado (compuerta G3 del programa)."),
    ("2 · Escalera [A]/[B]/desc. contenedor/[C]", "INPUTS C34–C37 — ratificar tras ver márgenes en hoja ESCALERA."),
    ("3 · EXW Tacna vs entregado", "Hoja FLETE — el flete por pack es 1.5–5% del precio público."),
    ("4 · Template facial", "Cajas/cupo 118 es BORRADOR no ratificado. Con retail S/25 el margen del comprador es ajustado; el escenario retail S/35 (segundo grid en FACIAL) lo convierte en producto estrella."),
    ("5 · Contenedor mixto", "Hoja MIXTO — higiénico + facial en el mismo cupo; resuelve la rotación del facial. Composición se ajusta con «cajas higiénico por cupo»."),
    ("", ""),
    ("RECOMENDACIÓN PRELIMINAR (a validar con estos números)", ""),
    ("· Precio publicado:", "EXW Tacna, un solo número por cupo. El mayorista de provincia tiene su transportista; empaquetar flete nos obliga a tarificar el peor caso."),
    ("· Cadenas (S1):", "cotizar ENTREGADO a su CD — las cadenas no compran EXW. Dos segmentos, dos incoterms internos."),
    ("· Punto de partida markup:", "40% — el comprador conserva ~25% sobre venta en higiénico; argumento de venta claro y defendible."),
    ("", ""),
    ("PREGUNTAS ABIERTAS (no las resuelve este libro)", ""),
    ("· Fiscal ZOFRATACNA/IGV:", "investigado 2026-07-16 — ver ZOFRATACNA.md en esta carpeta: la salida al resto del país es una importación (ad valorem + IGV/IPM 18% + percepción); dos estructuras posibles (Áladín nacionaliza vs comprador nacionaliza). Confirmar base del costo S/16 con agencia de aduana."),
    ("· Ruta/hub:", "el contenedor seed dice Qingdao→Callao; el almacén principal es Tacna. Confirmar ruta real antes de fijar copy del instrumento."),
    ("· Tarifas de flete:", "las de la hoja FLETE son PLACEHOLDERS — reemplazar con cotizaciones de transportista antes de publicar estimados."),
]
r = 3
for a, b in rows:
    ws.cell(row=r, column=1, value=a).font = Font(bold=bool(a and not a.startswith("·")), size=10)
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
widths(ws, {"A": 38, "B": 95})

# ───────────────────────────── INPUTS ─────────────────────────────
ws = wb.create_sheet("INPUTS")
ws["A1"] = "DRIVERS — todas las celdas amarillas son editables"
ws["A1"].font = f_title

sec(ws, 4, "GENERAL", 4)
driver(ws, 5, "Tipo de cambio (S/ por USD)", 3.75, NUM2)
driver(ws, 6, "IGV referencial", 0.18, PCT, "ver nota fiscal abajo")

sec(ws, 8, "PAPEL HIGIÉNICO — pack ×10 rollos", 4)
driver(ws, 9, "Costo por pack, almacén Tacna (S/)", 16, SOL, "dato Muaaz 2026-07-16")
driver(ws, 10, "Precio público objetivo por pack (S/)", 30, SOL)
calc(ws, 11, "Spread total por pack (S/)", "=C10-C9", SOL, "a repartir entre Wings y el comprador")
driver(ws, 12, "Packs por caja máster", 6, NUM)
driver(ws, 13, "Cajas por cupo", 94, NUM, "template 40HC ratificado v2")
driver(ws, 14, "Cupos por contenedor (40HC)", 10, NUM)
calc(ws, 15, "Packs por cupo", "=C12*C13", NUM)
calc(ws, 16, "Cajas comerciales por contenedor", "=C13*C14", NUM)
driver(ws, 17, "Cajas de holgura (stock propio)", 5, NUM)
driver(ws, 18, "m³ por caja", 0.0777, "0.0000", "dims SPSA 2020 — confirmar producción actual")
driver(ws, 19, "kg por caja", 9.7, NUM2)

sec(ws, 21, "PAPEL FACIAL — pack ×5 empaques (390 hojas) · TEMPLATE BORRADOR", 4)
driver(ws, 22, "Costo por pack ×5, almacén Tacna (S/)", 14, SOL, "dato Muaaz 2026-07-16")
driver(ws, 23, "Precio público objetivo por pack ×5 (S/)", 25, SOL, "= S/5 por empaque")
calc(ws, 24, "Spread total por pack (S/)", "=C23-C22", SOL)
driver(ws, 25, "Packs (×5) por caja máster", 9, NUM)
driver(ws, 26, "Cajas por cupo", 118, NUM, "BORRADOR — no ratificado (10×118=1.180 en 40HC)")
driver(ws, 27, "Cupos por contenedor", 10, NUM)
calc(ws, 28, "Packs por cupo", "=C25*C26", NUM)
driver(ws, 29, "m³ por caja", 0.059, "0.0000")
driver(ws, 30, "kg por caja", 9.7, NUM2)
driver(ws, 31, "Precio público facial — escenario alto (S/)", 35, SOL,
       "escenario pedido 2026-07-16; grid propio en hoja FACIAL")

sec(ws, 32, "ESCALERA (GTM §2.2) — drivers de decisión", 4)
driver(ws, 33, "Escenario base — markup Wings sobre costo", 0.40, PCT, "el dial principal: define el precio publicado")
driver(ws, 34, "[A] Descuento CUPO×3 (2–3 cupos)", 0.03, PCT)
driver(ws, 35, "[B] Descuento MEDIO (5 cupos)", 0.06, PCT)
driver(ws, 36, "Descuento CONTENEDOR (10 cupos)", 0.08, PCT)
driver(ws, 37, "[C] Crédito padrino (s/ 1 cupo, ciclo siguiente)", 0.02, PCT)

ws["A39"] = "NOTA FISCAL (investigado 2026-07-16, ver ZOFRATACNA.md): la salida de mercancía NO producida en zona"
ws["A40"] = "hacia el resto del país es una importación: ad valorem (banda probable 6%; TLC China puede dar 0% con"
ws["A41"] = "certificado de origen) + IGV+IPM 18% + percepción. Confirmar con agencia de aduana la base del costo S/16."
ws["A42"] = "NOTA DE DATOS: costos vigentes declarados por Muaaz 2026-07-16; sustituyen el costeo SPSA 2020 (muerto)."
for rr in (39, 40, 41, 42):
    ws.cell(row=rr, column=1).font = f_warn if rr < 42 else f_note
widths(ws, {"A": 44, "B": 4, "C": 14, "D": 52})

# ─────────────── grid builder for HIGIENICO / FACIAL ───────────────
HDRS = ["Markup s/ costo", "Precio S/ /pack", "Margen Wings S/ /pack", "Margen Wings % s/ precio",
        "Precio por cupo (S/)", "Precio por cupo (USD)", "Margen comprador S/ /pack",
        "Margen comprador % s/ venta", "Markup comprador s/ su costo",
        "Ganancia comprador por cupo (S/)", "Margen Wings por cupo (S/)",
        "Margen Wings por contenedor (S/)"]
FMTS = [PCT, SOL, SOL, PCT, SOL0, USD, SOL, PCT, PCT, SOL0, SOL0, SOL0]

def scenario_sheet(name, title, cost, retail, packs, cupos):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = f_title
    ws["A2"] = f"Costo y precio público en INPUTS ({cost} / {retail}); cambiarlos allí recalcula todo."
    ws["A2"].font = f_note
    for i, h in enumerate(HDRS, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = f_hdr; c.fill = fill_hdr; c.border = box
        c.alignment = Alignment(wrap_text=True, vertical="center")
    def grid_row(r, markup_expr, note=None):
        ws.cell(row=r, column=1, value=markup_expr).number_format = PCT
        ws.cell(row=r, column=2, value=f"=INPUTS!{cost}*(1+$A{r})")
        ws.cell(row=r, column=3, value=f"=B{r}-INPUTS!{cost}")
        ws.cell(row=r, column=4, value=f"=C{r}/B{r}")
        ws.cell(row=r, column=5, value=f"=B{r}*INPUTS!{packs}")
        ws.cell(row=r, column=6, value=f"=E{r}/INPUTS!$C$5")
        ws.cell(row=r, column=7, value=f"=INPUTS!{retail}-B{r}")
        ws.cell(row=r, column=8, value=f"=G{r}/INPUTS!{retail}")
        ws.cell(row=r, column=9, value=f"=G{r}/B{r}")
        ws.cell(row=r, column=10, value=f"=G{r}*INPUTS!{packs}")
        ws.cell(row=r, column=11, value=f"=C{r}*INPUTS!{packs}")
        ws.cell(row=r, column=12, value=f"=K{r}*INPUTS!{cupos}")
        for i, fmt in enumerate(FMTS, start=1):
            ws.cell(row=r, column=i).number_format = fmt
            ws.cell(row=r, column=i).border = box
        if note:
            ws.cell(row=r, column=13, value=note).font = f_note
    r = 5
    for m in (0.30, 0.35, 0.40, 0.45, 0.50):
        grid_row(r, m)
        r += 1
    # 50/50 split of the spread
    ws.cell(row=r, column=1, value=f"=B{r}/INPUTS!{cost}-1").number_format = PCT
    ws.cell(row=r, column=2, value=f"=INPUTS!{cost}+(INPUTS!{retail}-INPUTS!{cost})/2")
    for col, formula in [(3, f"=B{r}-INPUTS!{cost}"), (4, f"=C{r}/B{r}"),
                         (5, f"=B{r}*INPUTS!{packs}"), (6, f"=E{r}/INPUTS!$C$5"),
                         (7, f"=INPUTS!{retail}-B{r}"), (8, f"=G{r}/INPUTS!{retail}"),
                         (9, f"=G{r}/B{r}"), (10, f"=G{r}*INPUTS!{packs}"),
                         (11, f"=C{r}*INPUTS!{packs}"), (12, f"=K{r}*INPUTS!{cupos}")]:
        ws.cell(row=r, column=col, value=formula)
    for i, fmt in enumerate(FMTS, start=1):
        ws.cell(row=r, column=i).number_format = fmt
        ws.cell(row=r, column=i).border = box
    ws.cell(row=r, column=13, value="← reparto igual del spread").font = f_note
    r += 2
    ws.cell(row=r, column=1, value="Convención alternativa: margen sobre PRECIO DE VENTA (precio = costo ÷ (1−margen))").font = f_hdr
    r += 1
    hdr2 = r
    for i, h in enumerate(["Margen s/ precio"] + HDRS[1:], start=1):
        c = ws.cell(row=hdr2, column=i, value=h)
        c.font = f_hdr; c.fill = fill_hdr; c.border = box
        c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1
    for m in (0.30, 0.40, 0.50):
        ws.cell(row=r, column=1, value=m).number_format = PCT
        ws.cell(row=r, column=2, value=f"=INPUTS!{cost}/(1-$A{r})")
        for col, formula in [(3, f"=B{r}-INPUTS!{cost}"), (4, f"=C{r}/B{r}"),
                             (5, f"=B{r}*INPUTS!{packs}"), (6, f"=E{r}/INPUTS!$C$5"),
                             (7, f"=INPUTS!{retail}-B{r}"), (8, f"=G{r}/INPUTS!{retail}"),
                             (9, f"=G{r}/B{r}"), (10, f"=G{r}*INPUTS!{packs}"),
                             (11, f"=C{r}*INPUTS!{packs}"), (12, f"=K{r}*INPUTS!{cupos}")]:
            ws.cell(row=r, column=col, value=formula)
        for i, fmt in enumerate(FMTS, start=1):
            ws.cell(row=r, column=i).number_format = fmt
            ws.cell(row=r, column=i).border = box
        ws.cell(row=r, column=13,
                value=f'=IF(B{r}>=INPUTS!{retail},"⚠ supera el precio público — inviable","")')
        ws.cell(row=r, column=13).font = f_warn
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Rotación del comprador — semanas para vender 1 cupo").font = f_hdr
    r += 1
    ws.cell(row=r, column=1, value="Packs vendidos / semana")
    for i, rate in enumerate((20, 40, 60, 100), start=2):
        c = ws.cell(row=r, column=i, value=rate)
        c.fill = fill_edit; c.border = box; c.number_format = NUM
    r += 1
    ws.cell(row=r, column=1, value="Semanas para rotar el cupo")
    for i in range(2, 6):
        col = get_column_letter(i)
        c = ws.cell(row=r, column=i, value=f"=INPUTS!{packs}/{col}{r-1}")
        c.number_format = "0.0"; c.border = box
    ws.freeze_panes = "A5"
    widths(ws, {get_column_letter(i): 13 for i in range(1, 13)})
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["M"].width = 40
    return ws

scenario_sheet("HIGIENICO", "PAPEL HIGIÉNICO — escenarios de precio de cupo",
               "$C$9", "$C$10", "$C$15", "$C$14")
fac = scenario_sheet("FACIAL", "PAPEL FACIAL — escenarios de precio de cupo (template BORRADOR)",
                     "$C$22", "$C$23", "$C$28", "$C$27")
fac["A3"] = "⚠ Cajas/cupo = 118 no ratificado. Margen del comprador más ajustado que higiénico — ver columnas H–I."
fac["A3"].font = f_warn

# Second facial grid — retail escenario alto (INPUTS C31)
fac["A23"] = "ESCENARIO RETAIL ALTO — precio público S/ 35 (INPUTS C31)"
fac["A23"].font = f_title
fac["A24"] = "Mismo costo S/14; solo cambia el precio de venta del comprador — su margen es la columna que se mueve."
fac["A24"].font = f_note
for i, h in enumerate(HDRS, start=1):
    c = fac.cell(row=26, column=i, value=h)
    c.font = f_hdr; c.fill = fill_hdr; c.border = box
    c.alignment = Alignment(wrap_text=True, vertical="center")
for j, m in enumerate((0.30, 0.35, 0.40, 0.45, 0.50)):
    r = 27 + j
    fac.cell(row=r, column=1, value=m)
    fac.cell(row=r, column=2, value=f"=INPUTS!$C$22*(1+$A{r})")
    fac.cell(row=r, column=3, value=f"=B{r}-INPUTS!$C$22")
    fac.cell(row=r, column=4, value=f"=C{r}/B{r}")
    fac.cell(row=r, column=5, value=f"=B{r}*INPUTS!$C$28")
    fac.cell(row=r, column=6, value=f"=E{r}/INPUTS!$C$5")
    fac.cell(row=r, column=7, value=f"=INPUTS!$C$31-B{r}")
    fac.cell(row=r, column=8, value=f"=G{r}/INPUTS!$C$31")
    fac.cell(row=r, column=9, value=f"=G{r}/B{r}")
    fac.cell(row=r, column=10, value=f"=G{r}*INPUTS!$C$28")
    fac.cell(row=r, column=11, value=f"=C{r}*INPUTS!$C$28")
    fac.cell(row=r, column=12, value=f"=K{r}*INPUTS!$C$27")
    for i, fmt in enumerate(FMTS, start=1):
        fac.cell(row=r, column=i).number_format = fmt
        fac.cell(row=r, column=i).border = box
fac["M27"] = "← a S/35 retail el facial pasa de producto ajustado a producto estrella del comprador"
fac["M27"].font = f_note

# ───────────────────────────── MIXTO ─────────────────────────────
ws = wb.create_sheet("MIXTO")
ws["A1"] = "CONTENEDOR MIXTO 40HC — higiénico + facial en un mismo cupo"
ws["A1"].font = f_title
ws["A2"] = "⚠ PROPUESTA — template no ratificado. La math del servidor (SPEC §3.2) manda; el factor de estiba mixta es supuesto hasta tener plan de carga real."
ws["A2"].font = f_warn

sec(ws, 4, "GEOMETRÍA", 4)
driver(ws, 5, "Volumen interior 40HC (m³)", 76.4, NUM2)
driver(ws, 6, "Factor de utilización — estiba mixta", 0.92, PCT,
       "grids puros lograron 96,1% (hig.) / 91,7% (facial)")
calc(ws, 7, "m³ útiles del contenedor", "=C5*C6", NUM2)
calc(ws, 8, "m³ útiles por cupo (÷ cupos)", "=C7/INPUTS!$C$14", NUM2)
driver(ws, 9, "Cajas higiénico por cupo", 66, NUM, "el dial de composición")
calc(ws, 10, "m³ higiénico por cupo", "=C9*INPUTS!$C$18", NUM2)
calc(ws, 11, "Cajas facial por cupo (calculado)", "=FLOOR((C8-C10)/INPUTS!$C$29,1)", NUM)
calc(ws, 12, "m³ usados por cupo", "=C10+C11*INPUTS!$C$29", NUM2)
calc(ws, 13, "Holgura m³ por cupo", "=C8-C12", NUM2)
calc(ws, 14, "kg por cupo", "=C9*INPUTS!$C$19+C11*INPUTS!$C$30", NUM)
calc(ws, 15, "Composición del contenedor",
     '="10 cupos · "&C9*10&" cajas higiénico + "&C11*10&" cajas facial"', "@")

sec(ws, 17, "EL CUPO MIXTO — escenario base (markup INPUTS C33)", 4)
hdrs_m = ["Concepto", "Higiénico", "Facial", "Total cupo"]
for i, h in enumerate(hdrs_m, start=1):
    c = ws.cell(row=18, column=i, value=h)
    c.font = f_hdr; c.fill = fill_hdr; c.border = box
mrows = [
    ("Cajas por cupo", "=C9", "=C11", "=B19+C19", NUM),
    ("Packs por cupo", "=B19*INPUTS!$C$12", "=C19*INPUTS!$C$25", "=B20+C20", NUM),
    ("Precio S/ /pack", "=INPUTS!$C$9*(1+INPUTS!$C$33)", "=INPUTS!$C$22*(1+INPUTS!$C$33)", None, SOL),
    ("Valor por cupo (S/)", "=B20*B21", "=C20*C21", "=B22+C22", SOL0),
    ("Costo por cupo (S/)", "=B20*INPUTS!$C$9", "=C20*INPUTS!$C$22", "=B23+C23", SOL0),
    ("Margen Wings por cupo (S/)", "=B22-B23", "=C22-C23", "=B24+C24", SOL0),
    ("Margen comprador S/ /pack", "=INPUTS!$C$10-B21", "=INPUTS!$C$23-C21", None, SOL),
    ("Margen comprador % s/ venta (retail base)", "=B25/INPUTS!$C$10", "=C25/INPUTS!$C$23", None, PCT),
    ("Margen comprador facial % (retail S/35)", None, "=(INPUTS!$C$31-C21)/INPUTS!$C$31", None, PCT),
]
for j, (label, bh, cf, dt, fmt) in enumerate(mrows):
    r = 19 + j
    ws.cell(row=r, column=1, value=label)
    for col, val in ((2, bh), (3, cf), (4, dt)):
        if val is not None:
            cc = ws.cell(row=r, column=col, value=val)
            cc.number_format = fmt; cc.border = box

sec(ws, 29, "P&L CONTENEDOR MIXTO", 4)
calc(ws, 30, "Ingreso 10 cupos (S/)", "=D22*10", SOL0)
calc(ws, 31, "Costo 10 cupos (S/)", "=D23*10", SOL0)
calc(ws, 32, "Margen bruto (S/)", "=C30-C31", SOL0)
calc(ws, 33, "Margen bruto % s/ ingreso", "=C32/C30", PCT)
calc(ws, 34, "Margen bruto (USD)", "=C32/INPUTS!$C$5", USD)
ws["A36"] = "Por qué existe esta hoja: el cupo mixto resuelve la rotación del facial —"
ws["A37"] = "~288 packs de facial por cupo (vs 1.062 en cupo puro) y le da al comprador surtido de góndola completo."
for rr in (36, 37):
    ws.cell(row=rr, column=1).font = f_note
widths(ws, {"A": 42, "B": 14, "C": 14, "D": 14})

# ───────────────────────────── ESCALERA ─────────────────────────────
ws = wb.create_sheet("ESCALERA")
ws["A1"] = "ESCALERA DE BENEFICIOS — higiénico, escenario base (INPUTS C33)"
ws["A1"].font = f_title
ws["A3"] = "Precio base S/ /pack"
ws["B3"] = "=INPUTS!$C$9*(1+INPUTS!$C$33)"
ws["B3"].number_format = SOL
ws["A4"] = "Precio base por cupo"
ws["B4"] = "=B3*INPUTS!$C$15"
ws["B4"].number_format = SOL0
hdrs = ["Tier", "Cupos", "Descuento", "S/ /pack efectivo", "Precio por cupo (S/)",
        "Total pedido (S/)", "Total pedido (USD)", "Margen Wings S/ /pack",
        "Margen Wings % s/ costo", "Margen comprador S/ /pack", "Margen comprador % s/ venta"]
for i, h in enumerate(hdrs, start=1):
    c = ws.cell(row=6, column=i, value=h)
    c.font = f_hdr; c.fill = fill_hdr; c.border = box
    c.alignment = Alignment(wrap_text=True, vertical="center")
tiers = [("CUPO", 1, "0"), ("CUPO×3", 3, "=INPUTS!$C$34"),
         ("MEDIO", 5, "=INPUTS!$C$35"), ("CONTENEDOR", 10, "=INPUTS!$C$36")]
for j, (name, n, desc) in enumerate(tiers):
    r = 7 + j
    ws.cell(row=r, column=1, value=name).font = Font(bold=True)
    ws.cell(row=r, column=2, value=n)
    ws.cell(row=r, column=3, value=desc).number_format = PCT
    ws.cell(row=r, column=4, value=f"=$B$3*(1-C{r})").number_format = SOL
    ws.cell(row=r, column=5, value=f"=D{r}*INPUTS!$C$15").number_format = SOL0
    ws.cell(row=r, column=6, value=f"=E{r}*B{r}").number_format = SOL0
    ws.cell(row=r, column=7, value=f"=F{r}/INPUTS!$C$5").number_format = USD
    ws.cell(row=r, column=8, value=f"=D{r}-INPUTS!$C$9").number_format = SOL
    ws.cell(row=r, column=9, value=f"=H{r}/INPUTS!$C$9").number_format = PCT
    ws.cell(row=r, column=10, value=f"=INPUTS!$C$10-D{r}").number_format = SOL
    ws.cell(row=r, column=11, value=f"=J{r}/INPUTS!$C$10").number_format = PCT
    for i in range(1, 12):
        ws.cell(row=r, column=i).border = box
ws["A12"] = "Crédito padrino por referido CONFIRMADO (se aplica al ciclo siguiente)"
ws["B12"] = "=INPUTS!$C$37*$B$4"
ws["B12"].number_format = SOL
ws["A14"] = "Los mismos % de escalera aplican a facial cuando su template se ratifique."
ws["A14"].font = f_note
widths(ws, {get_column_letter(i): 14 for i in range(1, 12)})
ws.column_dimensions["A"].width = 18

# ───────────────────────────── FLETE ─────────────────────────────
ws = wb.create_sheet("FLETE")
ws["A1"] = "FLETE TACNA → DESTINO · por cupo de higiénico"
ws["A1"].font = f_title
ws["A2"] = "⚠ Tarifas S/ /m³ son PLACEHOLDERS EDITABLES — reemplazar con cotizaciones reales de transportista."
ws["A2"].font = f_warn
ws["A4"] = "m³ por cupo"
ws["B4"] = "=INPUTS!$C$13*INPUTS!$C$18"
ws["B4"].number_format = "0.00"
ws["A5"] = "kg por cupo"
ws["B5"] = "=INPUTS!$C$13*INPUTS!$C$19"
ws["B5"].number_format = NUM
ws["A6"] = "El papel cubica antes de pesar → la tarifa por m³ manda, no el kg."
ws["A6"].font = f_note
hdrs = ["Ciudad", "km aprox", "Tarifa S/ /m³ (EDITABLE)", "Flete por cupo (S/)",
        "Flete por pack (S/)", "% del precio público", "Cupo entregado (S/)"]
for i, h in enumerate(hdrs, start=1):
    c = ws.cell(row=8, column=i, value=h)
    c.font = f_hdr; c.fill = fill_hdr; c.border = box
    c.alignment = Alignment(wrap_text=True, vertical="center")
cities = [("Moquegua", 160, 35), ("Arequipa", 370, 45), ("Puno", 380, 50),
          ("Juliaca", 400, 52), ("Cusco", 790, 65), ("Lima", 1290, 85),
          ("Trujillo", 1850, 105), ("Piura", 2270, 120)]
for j, (city, km, rate) in enumerate(cities):
    r = 9 + j
    ws.cell(row=r, column=1, value=city)
    ws.cell(row=r, column=2, value=km).number_format = NUM
    c = ws.cell(row=r, column=3, value=rate)
    c.fill = fill_edit; c.number_format = SOL0
    ws.cell(row=r, column=4, value=f"=$B$4*C{r}").number_format = SOL0
    ws.cell(row=r, column=5, value=f"=D{r}/INPUTS!$C$15").number_format = SOL
    ws.cell(row=r, column=6, value=f"=E{r}/INPUTS!$C$10").number_format = PCT
    ws.cell(row=r, column=7, value=f"=ESCALERA!$B$4+D{r}").number_format = SOL0
    for i in range(1, 8):
        ws.cell(row=r, column=i).border = box
ws["A18"] = "Lectura: el flete por pack queda en céntimos → no justifica empaquetarlo en el precio del cupo."
ws["A19"] = "Recomendación S2 (mayoristas): precio EXW Tacna único + estimador de entrega transparente."
ws["A20"] = "Recomendación S1 (cadenas): cotizar entregado a su CD — las cadenas no compran EXW."
for rr in (18, 19, 20):
    ws.cell(row=rr, column=1).font = Font(bold=(rr > 18), size=9)
widths(ws, {"A": 14, "B": 10, "C": 20, "D": 16, "E": 15, "F": 16, "G": 17})

# ───────────────────────────── CONTENEDOR ─────────────────────────────
ws = wb.create_sheet("CONTENEDOR")
ws["A1"] = "P&L POR CONTENEDOR — higiénico 40HC, escenario base (INPUTS C33)"
ws["A1"].font = f_title
ws["A3"] = "Costo total (945 cajas incl. holgura)"
ws["B3"] = "=(INPUTS!$C$16+INPUTS!$C$17)*INPUTS!$C$12*INPUTS!$C$9"
ws["A4"] = "Ingreso 10 cupos sin descuento"
ws["B4"] = "=ESCALERA!$B$4*INPUTS!$C$14"
ws["A5"] = "Margen bruto (S/)"
ws["B5"] = "=B4-B3"
ws["A6"] = "Margen bruto % s/ ingreso"
ws["B6"] = "=B5/B4"
ws["B6"].number_format = PCT
ws["A7"] = "Margen bruto (USD)"
ws["B7"] = "=B5/INPUTS!$C$5"
ws["B7"].number_format = USD
for rr in (3, 4, 5):
    ws.cell(row=rr, column=2).number_format = SOL0
ws["A9"] = "ESCENARIOS DE MEZCLA DE TIERS"
ws["A9"].font = f_hdr
for i, h in enumerate(["Mezcla", "Ingreso (S/)", "Margen bruto (S/)", "% s/ ingreso"], start=1):
    c = ws.cell(row=10, column=i, value=h)
    c.font = f_hdr; c.fill = fill_hdr; c.border = box
mixes = [
    ("10 × CUPO (sin descuento)", "=ESCALERA!$E$7*10"),
    ("2×CUPO + 3×CUPO×3 + 5×MEDIO", "=2*ESCALERA!$E$7+3*ESCALERA!$E$8+5*ESCALERA!$E$9"),
    ("1 × CONTENEDOR dedicado", "=ESCALERA!$F$10"),
]
for j, (label, formula) in enumerate(mixes):
    r = 11 + j
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=formula).number_format = SOL0
    ws.cell(row=r, column=3, value=f"=B{r}-$B$3").number_format = SOL0
    ws.cell(row=r, column=4, value=f"=C{r}/B{r}").number_format = PCT
    for i in range(1, 5):
        ws.cell(row=r, column=i).border = box
ws["A16"] = "SENSIBILIDAD AL COSTO (precio publicado fijo, ingreso = mezcla 10×CUPO)"
ws["A16"].font = f_hdr
for i, h in enumerate(["Costo", "Costo total (S/)", "Margen bruto (S/)", "% s/ ingreso"], start=1):
    c = ws.cell(row=17, column=i, value=h)
    c.font = f_hdr; c.fill = fill_hdr; c.border = box
for j, (label, factor) in enumerate([("−10%", 0.9), ("base", 1.0), ("+10%", 1.1)]):
    r = 18 + j
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=f"=$B$3*{factor}").number_format = SOL0
    ws.cell(row=r, column=3, value=f"=$B$11-B{r}").number_format = SOL0
    ws.cell(row=r, column=4, value=f"=C{r}/$B$11").number_format = PCT
    for i in range(1, 5):
        ws.cell(row=r, column=i).border = box
widths(ws, {"A": 34, "B": 16, "C": 16, "D": 13})

wb.save(OUT)
print("saved", OUT)
