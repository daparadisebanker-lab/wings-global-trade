#!/usr/bin/env python3
"""Build an English-language Excel spec sheet for the Toyota Hilux Travo
Overland Plus 4TREX, translating the same data used in ficha.pdf (this
folder) into a clean workbook. Run:
  python3 build_specs_xlsx.py
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = HERE / "Hilux_Travo_Overland_Plus_4TREX_Specifications.xlsx"

DOC_NUMBER = "FT-WGT-2026-0826"
DOC_DATE = "26-08-2026"
MODEL_NAME = "Toyota Hilux Travo Overland Plus 4TREX"
MODEL_TRIM = "2.8L Diesel · 4WD · 6-Speed Automatic"

HERO_STATS = [
    ("Max. Power", "204 PS"),
    ("Max. Torque", "500 N·m"),
    ("Transmission", "6AT"),
    ("Drivetrain", "4WD · LHD"),
]

# (section index, section title, [(label, value), ...]) — translated from ficha.pdf
SECTIONS = [
    (1, "Identification", [
        ("Model", "Toyota Hilux Travo Overland Plus"),
        ("Version", "2.8L Diesel · 4WD · 6-Speed Automatic"),
        ("Body style", "Pickup, 4 doors, double cab"),
        ("Fuel type", "Diesel"),
        ("Manufacturer", "Toyota"),
        ("Steering position", "Left-hand drive (LHD) — confirmed by supplier"),
    ]),
    (2, "Dimensions & Weights", [
        ("Length × Width × Height (mm)", "5,320 × 1,885 × 1,865"),
        ("Wheelbase (mm)", "3,085"),
    ]),
    (3, "Engine & Transmission", [
        ("Engine model", "1GD-FTV (High)"),
        ("Engine type", "2.8L diesel, 4-cylinder, 16-valve DOHC, variable nozzle turbo with air-to-air intercooler"),
        ("Displacement (cc)", "2,755"),
        ("Max. power", "204 PS / 3,000–3,400 rpm"),
        ("Max. torque", "500 N·m / 1,600–2,800 rpm"),
        ("Transmission", "6-speed automatic"),
        ("Steering", "Electric power steering (EPS)"),
        ("Differential", "Rear locking differential"),
    ]),
    (4, "Chassis, Brakes & Tires", [
        ("Brakes (Front)", "Ventilated discs"),
        ("Brakes (Rear)", "Ventilated discs"),
        ("Suspension", "Front double wishbone / Rear rigid-axle leaf springs"),
        ("Tires / Wheels", "265/60 R18 alloy"),
    ]),
    (5, "Active Safety & Driver Assistance", [
        ("Toyota Safety Sense (TSS)", "Yes"),
        ("Anti-lock Braking System (ABS) with Electronic Brakeforce Distribution (EBD)", "Yes"),
        ("Vehicle Stability Control (VSC) with Brake Assist (BA)", "Yes"),
        ("Hill-start Assist Control (HAC)", "Yes"),
        ("Cruise control", "Dynamic Radar Cruise Control (adaptive)"),
        ("Blind Spot Monitor (BSM)", "Yes"),
        ("Rear Cross Traffic Alert (RCTA)", "Yes"),
        ("Pre-Collision System (PCS)", "Yes"),
        ("Lane Departure Alert (LDA)", "Yes"),
        ("Camera", "Panoramic View Monitor"),
        ("Parking sensors", "Front 4 / Rear 4"),
        ("Toyota Vehicle Security System (TVSS)", "Immobilizer + horn"),
    ]),
    (6, "Passive Safety", [
        ("SRS airbags", "Driver, driver knee, front passenger, side, curtain (7 total)"),
        ("Front seatbelts", "3-point ELR + pretensioner + force limiter"),
        ("Rear seatbelts", "3-point ELR x3"),
        ("Child restraint system", "ISOFIX x2 + tether anchor x2"),
    ]),
    (7, "Interior Equipment", [
        ("Seats (material)", "Synthetic leather"),
        ("Driver seat", "8-way power adjustment"),
        ("Front passenger seat", "4-way manual adjustment"),
        ("Rear seat", "60:40 split"),
        ("Steering wheel", "Electric power assist"),
        ("Steering wheel controls", "Audio, phone, display, voice control, cruise control, lane-keep assist"),
        ("Cup holders / bottle holders", "6 cup holders / 4 bottle holders"),
        ("Multi-information display (MID)", "12.3\" TFT"),
        ("Interior rearview mirror", "Auto-dimming"),
        ("Air conditioning", "Dual-zone automatic climate control"),
        ("Data communication module", "With remote function (coming soon)"),
        ("Hood lift assist", "Yes"),
        ("Audio system", "12.3\" display with 8 speakers"),
        ("Audio function", "Wireless Apple CarPlay and Android Auto"),
        ("12V outlet", "1"),
        ("Power windows", "Auto up/down with anti-pinch protection"),
        ("Door locks", "Speed-sensing"),
        ("Keyless entry", "Smart Entry"),
        ("Ignition", "Push Start button"),
    ]),
    (8, "Exterior Equipment & Lighting", [
        ("Headlights", "LED with auto-leveling"),
        ("Daytime running lights (DRL)", "LED"),
        ("Front fog lights", "LED"),
        ("Fender flares", "Included"),
        ("Front wipers", "Intermittent + interval adjustment"),
        ("Combination tail lights", "LED"),
        ("Exterior mirrors", "Power adjustment + power folding"),
    ]),
]

# ── Styling ─────────────────────────────────────────────────────────────
FONT_NAME = "Arial"
NAVY = "1F3864"
LIGHT_BAR = "F2F2F2"
LINE = "D9D9D9"
MUTED = "6B7280"

title_font = Font(name=FONT_NAME, size=20, bold=True, color=NAVY)
subtitle_font = Font(name=FONT_NAME, size=11, color=MUTED)
meta_font = Font(name=FONT_NAME, size=10, color=MUTED)
hero_label_font = Font(name=FONT_NAME, size=9, bold=True, color=MUTED)
hero_value_font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
section_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
section_idx_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
label_font = Font(name=FONT_NAME, size=10, bold=True)
value_font = Font(name=FONT_NAME, size=10)
footer_font = Font(name=FONT_NAME, size=9, color=MUTED, italic=True)

section_fill = PatternFill("solid", fgColor=NAVY)
hero_fill = PatternFill("solid", fgColor=LIGHT_BAR)
thin_bottom = Border(bottom=Side(style="thin", color=LINE))

wb = Workbook()
ws = wb.active
ws.title = "Specifications"
ws.sheet_view.showGridLines = False

ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 46
ws.column_dimensions["C"].width = 60

row = 1

def merge_row(r, first_col, last_col):
    ws.merge_cells(start_row=r, start_column=first_col, end_row=r, end_column=last_col)

# Title block
merge_row(row, 2, 3)
c = ws.cell(row=row, column=2, value="Technical Specification Sheet")
c.font = title_font
row += 1
merge_row(row, 2, 3)
c = ws.cell(row=row, column=2, value=DOC_NUMBER)
c.font = meta_font
row += 2

merge_row(row, 2, 3)
c = ws.cell(row=row, column=2, value=f"Prepared: {DOC_DATE}    |    Origin: China    |    Segment: 4x4 Pickup")
c.font = subtitle_font
row += 2

# Model hero
merge_row(row, 2, 3)
c = ws.cell(row=row, column=2, value=MODEL_NAME)
c.font = Font(name=FONT_NAME, size=15, bold=True, color=NAVY)
row += 1
merge_row(row, 2, 3)
c = ws.cell(row=row, column=2, value=MODEL_TRIM)
c.font = subtitle_font
row += 1

for label, value in HERO_STATS:
    ws.cell(row=row, column=2, value=label).font = hero_label_font
    ws.cell(row=row, column=2).fill = hero_fill
    ws.cell(row=row, column=3, value=value).font = hero_value_font
    ws.cell(row=row, column=3).fill = hero_fill
    row += 1
row += 1

# Column headers
ws.cell(row=row, column=2, value="Specification").font = Font(name=FONT_NAME, size=10, bold=True, color=MUTED)
ws.cell(row=row, column=3, value="Detail").font = Font(name=FONT_NAME, size=10, bold=True, color=MUTED)
row += 1

for idx, title, rows in SECTIONS:
    merge_row(row, 2, 3)
    c = ws.cell(row=row, column=2, value=f"{idx:02d}  {title.upper()}")
    c.font = section_font
    c.fill = section_fill
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 20
    ws.cell(row=row, column=3).fill = section_fill
    row += 1
    for label, value in rows:
        lc = ws.cell(row=row, column=2, value=label)
        lc.font = label_font
        lc.alignment = Alignment(wrap_text=True, vertical="top")
        lc.border = thin_bottom
        vc = ws.cell(row=row, column=3, value=value)
        vc.font = value_font
        vc.alignment = Alignment(wrap_text=True, vertical="top")
        vc.border = thin_bottom
        row += 1
    row += 1

# Footer / disclaimer
merge_row(row, 2, 3)
c = ws.cell(
    row=row, column=2,
    value=("The specifications above are provided for technical reference and may vary by "
           "production batch; confirmation against the physical unit prior to purchase is recommended."),
)
c.font = footer_font
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[row].height = 28
row += 2

merge_row(row, 2, 3)
c = ws.cell(row=row, column=2, value="WINGS GLOBAL TRADE")
c.font = Font(name=FONT_NAME, size=10, bold=True)
row += 1
merge_row(row, 2, 3)
c = ws.cell(row=row, column=2, value="Inquiries: importaciones@wingsglobaltrade.com   ·   Tel: +507 6025-07   ·   wingsglobaltrade.com")
c.font = meta_font

wb.save(OUT)
print(f"wrote {OUT}")
